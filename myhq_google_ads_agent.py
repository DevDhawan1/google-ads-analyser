#!/usr/bin/env python3
"""
myHQ Google Ads Analysis Agent
─────────────────────────────────────────────────────────────────────
Pulls live Google Ads campaign data and generates a VP-ready HTML
performance report. Mirrors the structure of myhq_ads_agent.py but
adapted for Google Ads API v17 via GAQL.

USAGE
  python myhq_google_ads_agent.py --developer-token <T> --client-id <ID>
    --client-secret <SECRET> --refresh-token <TOKEN> --customer-id <ID>
  python myhq_google_ads_agent.py --days 14 --out report.html

ENV VARS (alternative to flags)
  GOOGLE_ADS_DEVELOPER_TOKEN, GOOGLE_ADS_CLIENT_ID,
  GOOGLE_ADS_CLIENT_SECRET, GOOGLE_ADS_REFRESH_TOKEN,
  GOOGLE_ADS_CUSTOMER_ID

REQUIREMENTS
  pip install google-ads openpyxl pandas
─────────────────────────────────────────────────────────────────────
"""

import argparse
import os
import sys
import time
import webbrowser
from collections import defaultdict
from datetime import datetime

try:
    from google.ads.googleads.client import GoogleAdsClient as _GAClient
    from google.ads.googleads.errors import GoogleAdsException
except ImportError:
    print("\n  ERROR: 'google-ads' library not found.")
    print("  Fix:   pip install google-ads\n")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════
CONFIG = {
    "api_version": "v17",
    "default_days": 30,
    "brand": "myHQ",
    "thresholds": {
        "cpl_green":  200,
        "cpl_amber":  350,
        "cpl_red":    500,
        "ctr_green":  1.0,
        "ctr_amber":  0.6,
        "ctr_red":    0.5,
        "roas_green": 3.0,
        "roas_red":   1.5,
        "imp_share_red": 30.0,
        "min_spend":  3000,
    },
}

T = CONFIG["thresholds"]


# ═══════════════════════════════════════════════════════════════
# GOOGLE ADS CLIENT
# ═══════════════════════════════════════════════════════════════
class GoogleAdsClient:
    def __init__(self, developer_token, client_id, client_secret, refresh_token, customer_id):
        credentials = {
            "developer_token": developer_token,
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
            "use_proto_plus": True,
        }
        self.client = _GAClient.load_from_dict(credentials)
        self.customer_id = customer_id.replace("-", "")
        self.ga_service = self.client.get_service("GoogleAdsService")

    def _run_query(self, query, retries=3):
        for attempt in range(retries):
            try:
                stream = self.ga_service.search_stream(
                    customer_id=self.customer_id,
                    query=query,
                )
                rows = []
                for batch in stream:
                    for row in batch.results:
                        rows.append(row)
                return rows
            except GoogleAdsException as exc:
                for error in exc.failure.errors:
                    if error.error_code.quota_error.name == "RESOURCE_EXHAUSTED":
                        wait = 60 * (attempt + 1)
                        _log(f"  Rate limited — waiting {wait}s …")
                        time.sleep(wait)
                        break
                else:
                    raise RuntimeError(f"Google Ads API error: {exc}") from exc
        raise RuntimeError("Max retries exceeded")

    def _row_to_dict(self, row):
        """Flatten a proto-plus row into a plain dict."""
        m = row.metrics
        c = row.campaign
        ag = row.ad_group if hasattr(row, "ad_group") else None
        ad = row.ad_group_ad if hasattr(row, "ad_group_ad") else None
        segs = row.segments if hasattr(row, "segments") else None

        d = {
            "campaign_id":   str(c.id),
            "campaign_name": c.name,
            "cost_micros":   m.cost_micros,
            "impressions":   m.impressions,
            "clicks":        m.clicks,
            "ctr":           m.ctr,
            "average_cpc":   m.average_cpc,
            "average_cpm":   m.average_cpm,
            "conversions":   m.conversions,
            "conversions_value": m.conversions_value,
        }
        if hasattr(m, "search_impression_share"):
            d["search_impression_share"] = m.search_impression_share
        if ag:
            d["ad_group_id"]   = str(ag.id)
            d["ad_group_name"] = ag.name
        if ad:
            d["ad_id"]   = str(ad.ad.id)
            d["ad_name"] = ad.ad.name or f"Ad {ad.ad.id}"
            d["ad_type"] = ad.ad.type_.name if hasattr(ad.ad, "type_") else ""
        return d

    def campaign_insights(self, days):
        query = f"""
            SELECT
              campaign.id, campaign.name, campaign.status,
              metrics.cost_micros, metrics.impressions, metrics.clicks,
              metrics.ctr, metrics.average_cpc, metrics.average_cpm,
              metrics.conversions, metrics.conversions_value,
              metrics.cost_per_conversion, metrics.search_impression_share,
              metrics.search_rank_lost_impression_share
            FROM campaign
            WHERE campaign.status = 'ENABLED'
              AND segments.date DURING LAST_{days}_DAYS
        """
        return [self._row_to_dict(r) for r in self._run_query(query)]

    def adgroup_insights(self, days):
        query = f"""
            SELECT
              campaign.id, campaign.name,
              ad_group.id, ad_group.name, ad_group.status,
              metrics.cost_micros, metrics.impressions, metrics.clicks,
              metrics.ctr, metrics.average_cpm, metrics.average_cpc,
              metrics.conversions, metrics.conversions_value,
              metrics.cost_per_conversion, metrics.search_impression_share
            FROM ad_group
            WHERE campaign.status = 'ENABLED'
              AND ad_group.status = 'ENABLED'
              AND segments.date DURING LAST_{days}_DAYS
        """
        return [self._row_to_dict(r) for r in self._run_query(query)]

    def ad_insights(self, campaign_ids, days):
        ids_str = ", ".join(campaign_ids)
        query = f"""
            SELECT
              campaign.id, campaign.name,
              ad_group.id, ad_group.name,
              ad_group_ad.ad.id, ad_group_ad.ad.name,
              ad_group_ad.ad.type,
              metrics.cost_micros, metrics.impressions, metrics.clicks,
              metrics.ctr, metrics.average_cpm,
              metrics.conversions, metrics.conversions_value,
              metrics.cost_per_conversion
            FROM ad_group_ad
            WHERE campaign.id IN ({ids_str})
              AND ad_group_ad.status = 'ENABLED'
              AND segments.date DURING LAST_{days}_DAYS
        """
        return [self._row_to_dict(r) for r in self._run_query(query)]

    def account_info(self):
        query = """
            SELECT customer.id, customer.descriptive_name, customer.currency_code
            FROM customer
            LIMIT 1
        """
        rows = self._run_query(query)
        if rows:
            r = rows[0]
            c = r.customer
            return {
                "id":       str(c.id),
                "name":     c.descriptive_name,
                "currency": c.currency_code,
            }
        return {"id": self.customer_id, "name": "Unknown", "currency": "INR"}


# ═══════════════════════════════════════════════════════════════
# DATA HELPERS
# ═══════════════════════════════════════════════════════════════
def spend(r):       return (r.get("cost_micros", 0) or 0) / 1_000_000
def impressions(r): return int(r.get("impressions", 0) or 0)
def clicks(r):      return int(r.get("clicks", 0) or 0)
def ctr(r):         return float(r.get("ctr", 0) or 0) * 100
def avg_cpc(r):     return (r.get("average_cpc", 0) or 0) / 1_000_000
def cpm(r):         return (r.get("average_cpm", 0) or 0) / 1_000_000
def conversions(r): return float(r.get("conversions", 0) or 0)
def conv_value(r):  return float(r.get("conversions_value", 0) or 0)
def imp_share(r):   return float(r.get("search_impression_share", 0) or 0) * 100

def cpl(r):
    c = conversions(r); s = spend(r)
    return s / c if c > 0 and s > 0 else None

def roas(r):
    v = conv_value(r); s = spend(r)
    return v / s if v > 0 and s > 0 else None


# ═══════════════════════════════════════════════════════════════
# CLASSIFY
# ═══════════════════════════════════════════════════════════════
def classify(r):
    s = spend(r)
    if s < T["min_spend"]:
        return ("grey", "Too Early")

    cp = cpl(r);   ro = roas(r)
    ct = ctr(r);   ish = imp_share(r)

    if conversions(r) == 0 and conv_value(r) == 0:
        return ("yellow", "Watch — No Conv. Attr.")

    if ro is not None:
        if ro < T["roas_red"]:
            return ("red", "Underperforming")
        if ro >= T["roas_green"]:
            return ("green", "Performing")

    if cp is not None:
        if cp > T["cpl_red"] or ct < T["ctr_red"]:
            return ("red", "Underperforming")
        if cp <= T["cpl_green"] and ct >= T["ctr_green"]:
            return ("green", "Performing")

    if ish > 0 and ish < T["imp_share_red"]:
        return ("yellow", "Watch — Low Imp. Share")

    return ("yellow", "Watch")


# ═══════════════════════════════════════════════════════════════
# FORMATTING HELPERS
# ═══════════════════════════════════════════════════════════════
def inr(v, show_sym=True):
    if v is None or v != v:
        return "N/A"
    sym = "₹" if show_sym else ""
    v = int(round(v))
    s = str(abs(v))
    if len(s) > 3:
        last3 = s[-3:]
        rest  = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.append(rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.append(rest)
        s = ",".join(reversed(parts)) + "," + last3
    return f"{sym}{s}"

def pct(v):
    return f"{float(v):.2f}%" if v is not None else "N/A"

def xfmt(v):
    return f"{float(v):.2f}×" if v is not None else "N/A"

def tier_badge(tier, label):
    icons   = {"green": "🟢", "red": "🔴", "yellow": "🟡", "grey": "⚪"}
    classes = {"green": "tier-green", "red": "tier-red",
               "yellow": "tier-yellow", "grey": "tier-grey"}
    return f'<span class="tier {classes[tier]}">{icons[tier]} {label}</span>'

def flag(val_str, color):
    cls = {"red": "flag-red", "green": "flag-green", "amber": "flag-amber", "": ""}
    return f'<td class="number {cls.get(color,"")}">{val_str}</td>'

def cpl_color(v):
    if v is None: return ("N/A", "")
    if v <= T["cpl_green"]: return (inr(v), "green")
    if v <= T["cpl_amber"]: return (inr(v), "amber")
    return (inr(v), "red")

def ctr_color(v):
    v = float(v or 0)
    s = f"{v:.2f}%"
    if v >= T["ctr_green"]: return (s, "green")
    if v >= T["ctr_amber"]: return (s, "amber")
    return (s, "red")

def roas_color(v):
    if v is None: return ("N/A", "")
    s = f"{v:.2f}×"
    if v >= T["roas_green"]: return (s, "green")
    if v >= T["roas_red"]:   return (s, "amber")
    return (s, "red")

def imp_share_color(v):
    v = float(v or 0)
    if v <= 0: return ("N/A", "")
    s = f"{v:.1f}%"
    if v < T["imp_share_red"]: return (s, "red")
    if v < 50: return (s, "amber")
    return (s, "green")

def key_signal(r, tier):
    ct = ctr(r); cp = cpl(r); ro = roas(r); ish = imp_share(r)
    s  = spend(r)
    if tier == "grey":
        return "Too new — learning phase"
    if tier == "red":
        if cp: return f"CPL {inr(cp)} — {((cp/T['cpl_green'])-1)*100:.0f}% above target"
        if ro: return f"ROAS {ro:.2f}× — below {T['roas_red']}× minimum"
        return f"CTR {ct:.2f}% below threshold"
    if ro and ro >= T["roas_green"]:
        return f"{ro:.2f}× ROAS — {inr(conv_value(r))} attributed value"
    if cp and cp <= T["cpl_green"]:
        return f"{int(conversions(r)):,} conv. · CPL {inr(cp)}"
    if ish > 0 and ish < T["imp_share_red"]:
        return f"Imp. Share {ish:.1f}% — budget or quality issue"
    if ct > 0:
        return f"CTR {ct:.2f}% · {impressions(r):,} impressions"
    return "—"


# ═══════════════════════════════════════════════════════════════
# TERMINAL LOGGING
# ═══════════════════════════════════════════════════════════════
def _log(msg):
    print(msg, flush=True)

def _ok(msg=""):
    print(f"  ✔  {msg}" if msg else "  ✔", flush=True)


# ═══════════════════════════════════════════════════════════════
# AUTO QUICK WINS
# ═══════════════════════════════════════════════════════════════
def auto_quick_wins(campaign_rows, adgroup_rows, ad_rows_map, red_rows):
    wins = []

    # 1. Best ROAS — scale it
    scalable = [r for r in campaign_rows if roas(r) and roas(r) >= T["roas_green"]]
    scalable.sort(key=roas, reverse=True)
    if scalable:
        r = scalable[0]
        wins.append({
            "rank": 1, "timing": "today", "label": "Do today",
            "title": f"Scale {r['campaign_name']} budget 20–30%",
            "body": (f"{roas(r):.2f}× ROAS, {ctr(r):.2f}% CTR — "
                     "significant headroom. Increase daily budget incrementally and monitor CPC.")
        })

    # 2. Highest CPL red campaign — pause or restructure
    if red_rows:
        worst = max(red_rows, key=spend)
        cp = cpl(worst)
        wins.append({
            "rank": len(wins) + 1, "timing": "urgent", "label": "Do today",
            "title": f"Investigate {worst['campaign_name']} immediately",
            "body": (
                (f"CPL {inr(cp)} — {((cp/T['cpl_green'])-1)*100:.0f}% above ₹{T['cpl_green']:,} target. " if cp else "")
                + f"{inr(spend(worst))} spent this period. "
                + "Review ad copy, landing page, and audience targeting. Consider pausing until fixed."
            )
        })

    # 3. Ad group CPL gap inside red campaigns
    if adgroup_rows:
        by_campaign = defaultdict(list)
        for a in adgroup_rows:
            by_campaign[a.get("campaign_id")].append(a)
        for cid, adgroups in by_campaign.items():
            cpls = [(a, cpl(a)) for a in adgroups if cpl(a)]
            if len(cpls) >= 2:
                cpls.sort(key=lambda x: x[1])
                best_ag, best_v = cpls[0]
                worst_ag, worst_v = cpls[-1]
                if worst_v > best_v * 2.5:
                    wins.append({
                        "rank": len(wins) + 1, "timing": "confirmed", "label": "Data confirmed ✓",
                        "title": f"Cap spend on '{worst_ag.get('ad_group_name')}' ad group in {worst_ag.get('campaign_name')}",
                        "body": (f"Runs at {inr(worst_v)} CPL vs {inr(best_v)} for "
                                 f"'{best_ag.get('ad_group_name')}' — "
                                 f"a {worst_v/best_v:.1f}× gap. Shift budget to the lower-CPL ad group.")
                    })
                    break

    # 4. Creative consolidation for red campaigns
    all_ad_rows = [a for ads in ad_rows_map.values() for a in ads]
    if all_ad_rows:
        by_campaign = defaultdict(list)
        for a in all_ad_rows:
            by_campaign[a.get("campaign_id")].append(a)
        for cid, ads in by_campaign.items():
            agg = defaultdict(lambda: {"spend": 0, "conversions": 0})
            for a in ads:
                n = a.get("ad_name", "Unknown")
                agg[n]["spend"]       += spend(a)
                agg[n]["conversions"] += conversions(a)
            agg_list = [(n, v) for n, v in agg.items() if v["conversions"] > 0]
            if len(agg_list) >= 2:
                agg_list.sort(key=lambda x: x[1]["spend"] / x[1]["conversions"])
                best_n, best_v = agg_list[0]
                bad = [(n, v) for n, v in agg_list
                       if v["spend"] / v["conversions"] > best_v["spend"] / best_v["conversions"] * 2.5]
                if bad:
                    wasted = sum(v["spend"] for _, v in bad)
                    campaign_name = ads[0].get("campaign_name", cid)
                    wins.append({
                        "rank": len(wins) + 1, "timing": "confirmed", "label": "Data confirmed ✓",
                        "title": f"Pause {len(bad)} underperforming ad(s) in {campaign_name}",
                        "body": (f"<strong>{best_n}</strong> delivers the lowest CPL but {len(bad)} other "
                                 f"ad(s) run at 2.5×+ that cost, burning ~{inr(wasted)}. "
                                 "Pause the underperformers and consolidate budget on the winner.")
                    })
                    break

    # 5. Low impression share warning
    low_ish = [r for r in campaign_rows if 0 < imp_share(r) < T["imp_share_red"]]
    low_ish.sort(key=spend, reverse=True)
    if low_ish:
        r = low_ish[0]
        wins.append({
            "rank": len(wins) + 1, "timing": "week", "label": "This week",
            "title": f"Fix impression share for {r['campaign_name']}",
            "body": (f"Search Impression Share is {imp_share(r):.1f}% — "
                     "below 30% threshold. This is either a budget constraint or a Quality Score / "
                     "bid issue. Check Lost IS (Budget) vs Lost IS (Rank) in Google Ads UI.")
        })

    return wins[:5]


# ═══════════════════════════════════════════════════════════════
# HTML CSS + JS
# ═══════════════════════════════════════════════════════════════
_CSS = """
  *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
  :root{
    --ink:#0f172a;--ink-mid:#334155;--ink-soft:#64748b;--ink-mute:#94a3b8;
    --border:#e2e8f0;--bg:#f1f5f9;--white:#fff;
    --accent:#0ea5e9;--accent-dk:#0369a1;
    --green:#16a34a;--green-bg:#dcfce7;
    --amber:#d97706;--amber-bg:#fef3c7;
    --red:#dc2626;--red-bg:#fee2e2;
    --grey-bg:#f1f5f9;--radius:10px;
    --shadow:0 1px 3px rgba(0,0,0,.07),0 1px 2px rgba(0,0,0,.05);
  }
  html{scroll-behavior:smooth}
  body{font-family:'Segoe UI','Helvetica Neue',Arial,sans-serif;font-size:14px;
       line-height:1.6;color:var(--ink);background:var(--bg)}
  .page{max-width:1180px;margin:0 auto;padding:0 24px 60px}
  .hdr{background:var(--ink)}
  .hdr-inner{max-width:1180px;margin:0 auto;padding:32px 32px 0}
  .hdr-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:20px}
  .hdr-brand{font-size:11px;letter-spacing:3px;text-transform:uppercase;color:#64748b;margin-bottom:6px}
  .hdr-title{font-size:28px;font-weight:800;color:#fff;letter-spacing:-.4px}
  .hdr-sub{font-size:14px;color:#94a3b8;margin-top:6px}
  .hdr-meta{text-align:right;font-size:13px;color:#94a3b8;line-height:1.9}
  .hdr-meta strong{color:#cbd5e1}
  .confidential{display:inline-block;font-size:10px;font-weight:700;letter-spacing:1.5px;
    text-transform:uppercase;border:1px solid #334155;color:#64748b;
    padding:3px 10px;border-radius:4px;margin-top:8px}
  .kpi-strip{display:grid;grid-template-columns:repeat(5,1fr);gap:1px;
    background:#1e293b;border-top:1px solid #1e293b;max-width:1180px;margin:0 auto}
  .kpi-card{background:#1e293b;padding:18px 24px}
  .kpi-val{font-size:26px;font-weight:800;color:#fff;letter-spacing:-.5px;line-height:1.1}
  .kpi-val.good{color:#4ade80}.kpi-val.warn{color:#fbbf24}.kpi-val.bad{color:#f87171}
  .kpi-lbl{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-top:4px}
  .tab-nav-wrap{background:#fff;border-bottom:2px solid var(--border);
    position:sticky;top:0;z-index:200;box-shadow:0 2px 8px rgba(0,0,0,.06)}
  .tab-nav{max-width:1180px;margin:0 auto;padding:0 32px;display:flex;overflow-x:auto}
  .tab-btn{padding:15px 22px;font-size:13px;font-weight:600;color:var(--ink-soft);
    border:none;background:none;cursor:pointer;border-bottom:3px solid transparent;
    margin-bottom:-2px;white-space:nowrap;transition:color .15s,border-color .15s}
  .tab-btn:hover{color:var(--ink)}
  .tab-btn.active{color:var(--ink);border-bottom-color:var(--accent)}
  .tab-badge{display:inline-block;font-size:10px;font-weight:700;
    padding:1px 7px;border-radius:10px;margin-left:6px;
    background:var(--red-bg);color:var(--red)}
  .tab-panel{display:none;padding-top:36px}
  .tab-panel.active{display:block}
  .eyebrow{font-size:11px;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;
    color:var(--accent);margin-bottom:6px;display:block}
  h2{font-size:22px;font-weight:800;color:var(--ink);margin-bottom:6px;letter-spacing:-.3px}
  h3{font-size:16px;font-weight:700;color:var(--ink);margin:32px 0 10px}
  p{color:var(--ink-mid);margin-bottom:10px;line-height:1.7}
  .card{background:#fff;border:1px solid var(--border);border-radius:var(--radius);
    padding:24px;box-shadow:var(--shadow);margin-bottom:20px}
  .table-wrap{background:#fff;border:1px solid var(--border);border-radius:var(--radius);
    overflow:hidden;box-shadow:var(--shadow);margin:14px 0 24px}
  table{width:100%;border-collapse:collapse;font-size:13px}
  thead tr{background:var(--ink);color:#fff}
  thead th{padding:12px 14px;text-align:left;font-weight:600;font-size:12px;
    letter-spacing:.4px;white-space:nowrap}
  tbody tr:nth-child(even){background:#fafbfc}
  tbody tr:hover{background:#f0f7ff;transition:background .1s}
  tbody td{padding:11px 14px;border-bottom:1px solid var(--border);vertical-align:middle}
  tbody tr:last-child td{border-bottom:none}
  td.number{text-align:right;font-feature-settings:"tnum"}
  td.bold{font-weight:600}
  .tier{display:inline-flex;align-items:center;gap:4px;padding:3px 11px;
    border-radius:20px;font-size:12px;font-weight:600;white-space:nowrap}
  .tier-green{background:var(--green-bg);color:#15803d}
  .tier-yellow{background:var(--amber-bg);color:#92400e}
  .tier-red{background:var(--red-bg);color:var(--red)}
  .tier-grey{background:var(--grey-bg);color:var(--ink-mid)}
  .flag-red{color:var(--red);font-weight:700}
  .flag-green{color:var(--green);font-weight:700}
  .flag-amber{color:var(--amber);font-weight:700}
  .alert-block{background:#fff8f8;border:1px solid #fca5a5;border-left:4px solid var(--red);
    border-radius:0 var(--radius) var(--radius) 0;padding:20px 24px;margin:20px 0}
  .alert-block h3{color:var(--red);margin:0 0 12px;font-size:15px}
  .alert-section-label{font-size:11px;font-weight:700;text-transform:uppercase;
    letter-spacing:1px;color:var(--red);margin:14px 0 4px}
  .alert-block ul{margin:8px 0 8px 20px}
  .alert-block li{line-height:1.9;color:var(--ink-mid)}
  .quick-wins{list-style:none}
  .quick-wins li{display:flex;gap:16px;padding:16px 0;
    border-bottom:1px solid var(--border);align-items:flex-start}
  .quick-wins li:last-child{border-bottom:none}
  .win-rank{flex-shrink:0;width:30px;height:30px;background:var(--ink);color:#fff;
    border-radius:50%;display:flex;align-items:center;justify-content:center;
    font-size:13px;font-weight:700;margin-top:2px}
  .win-body{flex:1}
  .win-body strong{display:block;font-size:14px;color:var(--ink);margin-bottom:4px}
  .win-body p{font-size:13px;margin:0}
  .win-timing{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;
    padding:2px 9px;border-radius:10px;background:#dbeafe;color:#1d4ed8;
    margin:4px 0 6px;display:inline-block}
  .win-timing.confirmed{background:var(--green-bg);color:#15803d}
  .win-timing.urgent{background:var(--red-bg);color:var(--red)}
  .win-timing.week{background:#f0fdf4;color:#15803d}
  .report-footer{margin-top:48px;padding-top:16px;border-top:1px solid var(--border);
    font-size:11px;color:var(--ink-mute);display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px}
  @media print{
    body{background:#fff}
    .tab-nav-wrap{display:none}
    .tab-panel{display:block!important;page-break-before:always}
    .tab-panel:first-of-type{page-break-before:avoid}
    .alert-block,.card,.table-wrap{break-inside:avoid}
    .page{padding:0 16px}
  }
"""

_JS = """
  function showTab(id,btn){
    document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
    document.getElementById('tab-'+id).classList.add('active');
    btn.classList.add('active');
    const nav=document.getElementById('tabNav');
    window.scrollTo({top:nav.getBoundingClientRect().top+window.scrollY-4,behavior:'smooth'});
  }
"""


# ═══════════════════════════════════════════════════════════════
# HTML SECTION BUILDERS
# ═══════════════════════════════════════════════════════════════
def _triage_row(idx, r, tier, label):
    cp = cpl(r); ct = ctr(r); ro = roas(r); ish = imp_share(r)

    if ro is not None:
        cv, cc = roas_color(ro)
        metric_str = cv
    elif cp is not None:
        cv, cc = cpl_color(cp)
        metric_str = cv
    else:
        cc = ""; metric_str = "N/A"

    ctr_s, ctr_c = ctr_color(ct)
    ish_s, ish_c = imp_share_color(ish)

    row_style = ' style="background:#fff0f0;"' if tier == "red" else ""
    return (
        f'<tr{row_style}>'
        f'<td style="color:var(--ink-mute);">{idx}</td>'
        f'<td class="bold">{r.get("campaign_name","—")}</td>'
        f'<td class="number">{inr(spend(r))}</td>'
        f'<td class="number">{int(conversions(r)):,}</td>'
        f'{flag(metric_str, cc)}'
        f'{flag(ctr_s, ctr_c)}'
        f'{flag(ish_s, ish_c)}'
        f'<td>{key_signal(r, tier)}</td>'
        f'<td>{tier_badge(tier, label)}</td>'
        f'</tr>\n'
    )

def build_triage_table(rows):
    header = (
        '<div class="table-wrap"><table>'
        '<thead><tr>'
        '<th style="width:40px">#</th>'
        '<th>Campaign</th>'
        '<th class="number">Spend</th>'
        '<th class="number">Conversions</th>'
        '<th class="number">CPL / ROAS</th>'
        '<th class="number">CTR</th>'
        '<th class="number">Imp. Share</th>'
        '<th>Key Signal</th>'
        '<th>Status</th>'
        '</tr></thead><tbody>\n'
    )
    body = ""
    for idx, r in enumerate(rows, 1):
        tier, label = classify(r)
        body += _triage_row(idx, r, tier, label)
    return header + body + '</tbody></table></div>\n'

def build_alert_block(r, adgroup_rows_for_campaign, ad_rows_for_campaign):
    name = r.get("campaign_name", "Unknown")
    s    = spend(r); cp = cpl(r); ro = roas(r); ct = ctr(r); ish = imp_share(r)

    html = [f'<div class="alert-block"><h3>🔴 {name}</h3>']
    html.append('<div class="alert-section-label">What\'s wrong</div>')

    issues = []
    if cp and cp > T["cpl_red"]:
        issues.append(f"CPL of <strong>{inr(cp)}</strong> is {((cp/T['cpl_green'])-1)*100:.0f}% above the ₹{T['cpl_green']:,} target.")
    if ro and ro < T["roas_red"]:
        issues.append(f"ROAS of <strong>{ro:.2f}×</strong> is below the minimum viable threshold of {T['roas_red']}×.")
    if ct < T["ctr_red"]:
        issues.append(f"CTR of <strong>{ct:.2f}%</strong> is below the 0.5% threshold — ad copy or targeting needs work.")
    if ish > 0 and ish < T["imp_share_red"]:
        issues.append(f"Search Impression Share of <strong>{ish:.1f}%</strong> is below 30% — "
                      "possible budget cap or low Quality Score causing missed auctions.")
    if not issues:
        issues.append(f"Performance below account benchmarks. Spend this period: <strong>{inr(s)}</strong>.")

    html.append('<p>' + ' '.join(issues) + '</p>')

    # Ad group gap signal
    if adgroup_rows_for_campaign:
        cpls = [(a, cpl(a)) for a in adgroup_rows_for_campaign if cpl(a)]
        if cpls:
            cpls.sort(key=lambda x: x[1], reverse=True)
            worst_a, worst_v = cpls[0]
            best_a, best_v   = cpls[-1]
            if worst_v > best_v * 1.8:
                html.append('<div class="alert-section-label">Ad Group Finding</div>')
                html.append(
                    f'<p><strong>{worst_a.get("ad_group_name")}</strong> runs at {inr(worst_v)} CPL vs '
                    f'{inr(best_v)} CPL for <strong>{best_a.get("ad_group_name")}</strong> — '
                    f'a {worst_v/best_v:.1f}× gap. Reduce its budget by 50–60% and shift to the '
                    f'best-performing ad group. See Ad Group Breakdown tab for detail.</p>'
                )

    # Creative finding
    if ad_rows_for_campaign:
        agg = defaultdict(lambda: {"spend": 0, "conversions": 0})
        for a in ad_rows_for_campaign:
            n = a.get("ad_name", "Unknown")
            agg[n]["spend"]       += spend(a)
            agg[n]["conversions"] += conversions(a)
        agg_list = [(n, v) for n, v in agg.items() if v["conversions"] > 0]
        if len(agg_list) >= 2:
            agg_list.sort(key=lambda x: x[1]["spend"] / x[1]["conversions"])
            best_n, best_v = agg_list[0]
            bad = [(n, v) for n, v in agg_list
                   if v["spend"] / v["conversions"] > best_v["spend"] / best_v["conversions"] * 2]
            if bad:
                wasted = sum(v["spend"] for _, v in bad)
                html.append('<div class="alert-section-label">Ad Finding</div>')
                html.append(
                    f'<p><strong>{best_n}</strong> is the winning ad at '
                    f'{inr(best_v["spend"]/best_v["conversions"])} CPL. '
                    f'{len(bad)} other ad(s) run at 2×+ that cost, '
                    f'burning ~{inr(wasted)} in misallocated spend. '
                    f'Pause underperformers. See Ad Analysis tab.</p>'
                )

    html.append('</div>\n')
    return ''.join(html)

def build_adgroup_section(campaign_name, adgroup_rows, tier):
    if not adgroup_rows:
        return ""
    adgroup_rows = sorted(adgroup_rows, key=spend, reverse=True)
    total_s = sum(spend(a) for a in adgroup_rows)

    html = [f'<h3>{campaign_name} &nbsp;{tier_badge(tier[0], tier[1])} &nbsp;— Ad Group Breakdown</h3>']
    html.append('<div class="table-wrap"><table>')
    html.append('<thead><tr>'
                '<th>Ad Group</th>'
                '<th class="number">Spend</th>'
                '<th class="number">Budget %</th>'
                '<th class="number">Conversions</th>'
                '<th class="number">CPL</th>'
                '<th class="number">ROAS</th>'
                '<th class="number">CTR</th>'
                '<th class="number">Imp. Share</th>'
                '<th>Status</th>'
                '</tr></thead><tbody>\n')

    for a in adgroup_rows:
        s   = spend(a)
        cp  = cpl(a); ro = roas(a)
        ct  = ctr(a); ish = imp_share(a)
        budget_pct = f"{(s/total_s*100):.0f}%" if total_s > 0 else "—"

        cp_s, cp_c = cpl_color(cp) if cp else ("—", "")
        ro_s, ro_c = roas_color(ro) if ro else ("—", "")
        ctr_s, ctr_c = ctr_color(ct)
        ish_s, ish_c = imp_share_color(ish)

        metric_s = ro_s if ro else cp_s
        metric_c = ro_c if ro else cp_c

        if cp and cp > T["cpl_red"]:
            status = tier_badge("red", "Cap Spend")
        elif cp and cp <= T["cpl_green"] and ct >= T["ctr_green"]:
            status = tier_badge("green", "Scale")
        elif ish > 0 and ish < T["imp_share_red"]:
            status = tier_badge("yellow", "Watch")
        else:
            status = tier_badge("yellow", "Hold")

        html.append(
            f'<tr>'
            f'<td class="bold">{a.get("ad_group_name","—")}</td>'
            f'<td class="number">{inr(s)}</td>'
            f'<td class="number">{budget_pct}</td>'
            f'<td class="number">{int(conversions(a)):,}</td>'
            f'{flag(metric_s, metric_c)}'
            f'{flag(ro_s if ro else "—", ro_c)}'
            f'{flag(ctr_s, ctr_c)}'
            f'{flag(ish_s, ish_c)}'
            f'<td>{status}</td>'
            f'</tr>\n'
        )

    html.append('</tbody></table></div>\n')
    return ''.join(html)

def build_ad_section(campaign_name, ad_rows):
    if not ad_rows:
        return ""

    adgroup_groups = defaultdict(list)
    for a in ad_rows:
        adgroup_groups[a.get("ad_group_name", "Unknown")].append(a)

    html = [f'<h3>{campaign_name} — Ad Performance</h3>']

    for adgroup_name, ads in adgroup_groups.items():
        html.append(
            f'<p style="margin:16px 0 6px;font-weight:700;color:#1e2d3d;font-size:13px;">'
            f'📦 {adgroup_name}</p>'
        )
        html.append('<div class="table-wrap"><table>')
        html.append('<thead><tr>'
                    '<th>Ad Name</th>'
                    '<th>Ad Type</th>'
                    '<th class="number">Spend</th>'
                    '<th class="number">Impressions</th>'
                    '<th class="number">CPM</th>'
                    '<th class="number">CTR</th>'
                    '<th class="number">Conversions</th>'
                    '<th class="number">CPL</th>'
                    '<th>Verdict</th>'
                    '</tr></thead><tbody>\n')

        cpls = {a.get("ad_name", "?"): cpl(a) for a in ads if cpl(a)}
        best = min(cpls, key=cpls.get) if cpls else None

        for a in sorted(ads, key=spend, reverse=True):
            ad_name  = a.get("ad_name", "Unknown")
            ad_type  = a.get("ad_type", "").replace("_", " ").title()
            s = spend(a); imp = impressions(a)
            cp = cpl(a); ct = ctr(a); cm = cpm(a)

            cp_s, cp_c = cpl_color(cp) if cp else ("No conv.", "red")
            ctr_s, ctr_c = ctr_color(ct)

            is_best    = (ad_name == best)
            row_style  = ' style="background:#dcfce7;"' if is_best else ""

            if is_best:
                verdict = tier_badge("green", "Keep + Scale ⭐")
            elif cp and cp > T["cpl_amber"]:
                verdict = tier_badge("red", "Pause ❌")
            else:
                verdict = tier_badge("yellow", "Test further 🔍")

            html.append(
                f'<tr{row_style}>'
                f'<td class="bold">{ad_name}</td>'
                f'<td style="color:var(--ink-soft);font-size:12px">{ad_type}</td>'
                f'<td class="number">{inr(s)}</td>'
                f'<td class="number">{imp:,}</td>'
                f'<td class="number">{inr(cm) if cm else "—"}</td>'
                f'{flag(ctr_s, ctr_c)}'
                f'<td class="number">{int(conversions(a)):,}</td>'
                f'{flag(cp_s, cp_c)}'
                f'<td>{verdict}</td>'
                f'</tr>\n'
            )

        html.append('</tbody></table></div>\n')

    return ''.join(html)

def build_quick_wins_html(wins):
    if not wins:
        return "<p>No quick wins generated — all campaigns are performing within targets.</p>"
    items = []
    for w in wins:
        timing_class = w.get("timing", "week")
        label = w.get("label", "This week")
        items.append(
            f'<li>'
            f'<div class="win-rank">{w["rank"]}</div>'
            f'<div class="win-body">'
            f'<strong>{w["title"]}</strong>'
            f'<span class="win-timing {timing_class}">{label}</span>'
            f'<p>{w["body"]}</p>'
            f'</div></li>\n'
        )
    return f'<div class="card"><ul class="quick-wins">{"".join(items)}</ul></div>\n'


# ═══════════════════════════════════════════════════════════════
# ASSEMBLE FULL HTML
# ═══════════════════════════════════════════════════════════════
def generate_html(account, campaign_rows, adgroup_rows, ad_rows_map, days, generated_at):
    brand      = CONFIG["brand"]
    acct_name  = account.get("name", brand)
    acct_id    = account.get("id", "")
    currency   = account.get("currency", "INR")
    period_end = generated_at.strftime("%d %b %Y")

    classified = [(r, *classify(r)) for r in campaign_rows]
    red_rows   = [r for r, tier, _ in classified if tier == "red"]
    n_red      = len(red_rows)

    total_spend_v  = sum(spend(r) for r in campaign_rows)
    total_conv     = sum(conversions(r) for r in campaign_rows)
    best_roas_v    = max((roas(r) for r in campaign_rows if roas(r)), default=None)
    best_cpl_v     = min((cpl(r)  for r in campaign_rows if cpl(r)),  default=None)
    avg_ish        = (sum(imp_share(r) for r in campaign_rows if imp_share(r) > 0)
                      / max(sum(1 for r in campaign_rows if imp_share(r) > 0), 1))

    adgroup_by_cid = defaultdict(list)
    for a in adgroup_rows:
        adgroup_by_cid[a.get("campaign_id")].append(a)

    wins = auto_quick_wins(campaign_rows, adgroup_rows, ad_rows_map, red_rows)

    H = []
    H.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{acct_name} — Google Ads Analysis · {period_end}</title>
<style>{_CSS}</style>
</head>
<body>
""")

    # Header + KPI strip
    H.append(f"""
<div class="hdr">
<div class="hdr-inner">
  <div class="hdr-top">
    <div>
      <div class="hdr-brand">{brand} · Performance Marketing</div>
      <div class="hdr-title">Google Ads — Performance Review</div>
      <div class="hdr-sub">Last {days} Days · {len(campaign_rows)} Active Campaigns · {period_end}</div>
    </div>
    <div class="hdr-meta">
      <strong>Account:</strong> {acct_name} ({acct_id})<br>
      <strong>Currency:</strong> {currency} &nbsp;·&nbsp;
      <strong>API:</strong> Google Ads {CONFIG['api_version']}<br>
      <strong>Generated:</strong> {generated_at.strftime('%d %b %Y %H:%M')}
      <div class="confidential">Confidential — Internal Use Only</div>
    </div>
  </div>
</div>
<div class="kpi-strip">
  <div class="kpi-card">
    <div class="kpi-val">{inr(total_spend_v)}</div>
    <div class="kpi-lbl">{days}-Day Total Spend</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val">{int(total_conv):,}</div>
    <div class="kpi-lbl">Total Conversions</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val good">{inr(best_cpl_v) if best_cpl_v else "N/A"}</div>
    <div class="kpi-lbl">Best CPL</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val good">{f"{best_roas_v:.2f}×" if best_roas_v else "N/A"}</div>
    <div class="kpi-lbl">Best ROAS</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val {"warn" if avg_ish < 50 else "good"}">{avg_ish:.1f}%</div>
    <div class="kpi-lbl">Avg Imp. Share</div>
  </div>
</div>
</div>
""")

    # Tab nav
    H.append(f"""
<div class="tab-nav-wrap" id="tabNav">
<div class="tab-nav">
  <button class="tab-btn active" onclick="showTab('overview',this)">📊 Overview</button>
  <button class="tab-btn" onclick="showTab('alerts',this)">⚠️ Alerts &amp; Actions <span class="tab-badge">{n_red}</span></button>
  <button class="tab-btn" onclick="showTab('adgroup',this)">🗺 Ad Group Breakdown</button>
  <button class="tab-btn" onclick="showTab('ads',this)">🎯 Ad Analysis</button>
</div>
</div>
<div class="page">
""")

    # Tab 1 — Overview
    H.append('<div class="tab-panel active" id="tab-overview">')
    H.append('<span class="eyebrow">Campaign Overview</span>')
    H.append(f'<h2>Campaign Triage <small>{len(campaign_rows)} active campaigns</small></h2>')
    H.append('<p>Row colours: 🟢 Performing · 🟡 Watch · 🔴 Underperforming · ⚪ Too Early</p>')
    H.append(build_triage_table(campaign_rows))
    H.append('</div>\n')

    # Tab 2 — Alerts & Actions
    H.append(f'<div class="tab-panel" id="tab-alerts">')
    H.append('<span class="eyebrow">Alerts &amp; Actions</span>')
    H.append(f'<h2>Underperforming Campaigns <small>{n_red} flagged</small></h2>')
    if not red_rows:
        H.append('<div class="card"><p>No underperforming campaigns in this period.</p></div>')
    else:
        for r in red_rows:
            cid = r.get("campaign_id")
            H.append(build_alert_block(r, adgroup_by_cid.get(cid, []), ad_rows_map.get(cid, [])))

    H.append('<h3>Quick Wins</h3>')
    H.append(build_quick_wins_html(wins))
    H.append('</div>\n')

    # Tab 3 — Ad Group Breakdown
    H.append('<div class="tab-panel" id="tab-adgroup">')
    H.append('<span class="eyebrow">Ad Group Breakdown</span>')
    H.append('<h2>All Active Ad Groups</h2>')
    for r, tier, label in classified:
        cid  = r.get("campaign_id")
        name = r.get("campaign_name", "Unknown")
        ags  = adgroup_by_cid.get(cid, [])
        H.append(build_adgroup_section(name, ags, (tier, label)))
    H.append('</div>\n')

    # Tab 4 — Ad Analysis
    H.append('<div class="tab-panel" id="tab-ads">')
    H.append('<span class="eyebrow">Ad Analysis</span>')
    H.append('<h2>Ad Performance — Underperforming Campaigns</h2>')
    if not ad_rows_map:
        H.append('<div class="card"><p>No underperforming campaigns — ad-level analysis skipped.</p></div>')
    else:
        for r, tier, label in classified:
            cid = r.get("campaign_id")
            if cid in ad_rows_map:
                H.append(build_ad_section(r.get("campaign_name", cid), ad_rows_map[cid]))
    H.append('</div>\n')

    H.append(f"""
<div class="report-footer">
  <span>{acct_name} · Google Ads Performance Review · {period_end} · Confidential</span>
  <span>Google Ads API {CONFIG['api_version']} · Read-only · Account {acct_id} · Generated {generated_at.strftime('%d %b %Y %H:%M')}</span>
</div>
</div><!-- /page -->
<script>{_JS}</script>
</body>
</html>
""")

    return ''.join(H)


# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════
def generate_excel(account, campaign_rows, adgroup_rows, ad_rows_map, days, generated_at, out_path):
    if not OPENPYXL_OK:
        _log("  ⚠  openpyxl not found — skipping Excel export. Fix: pip install openpyxl")
        return None

    NAVY    = "1E2D3D"
    WHITE   = "FFFFFF"
    C_GREEN = "DCFCE7"
    C_AMBER = "FEF9C3"
    C_RED   = "FFF0F0"
    C_GREY  = "F3F4F6"

    def _fill(hex_c):       return PatternFill("solid", fgColor=hex_c)
    def _font(bold=False, size=10, color="000000", italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _tier_fill(tier):
        return {"green": _fill(C_GREEN), "yellow": _fill(C_AMBER),
                "red": _fill(C_RED), "grey": _fill(C_GREY)}.get(tier, None)
    def _cpl_fill(v):
        if v is None: return None
        if v <= T["cpl_green"]: return _fill(C_GREEN)
        if v <= T["cpl_amber"]: return _fill(C_AMBER)
        return _fill(C_RED)
    def _n(v, decimals=0):
        if v is None: return None
        try:
            return round(float(v), decimals) if decimals else int(round(float(v)))
        except (TypeError, ValueError):
            return None

    def _write_header(ws, cols, row=1, height=22):
        for ci, (text, width) in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=text)
            c.fill = _fill(NAVY); c.font = _font(bold=True, size=10, color=WHITE)
            c.alignment = _align("center")
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[row].height = height
        return len(cols)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    acct_name  = account.get("name", CONFIG["brand"])
    acct_id    = account.get("id", "")
    currency   = account.get("currency", "INR")
    classified = [(r, *classify(r)) for r in campaign_rows]
    red_rows   = [r for r, t, _ in classified if t == "red"]

    total_spend_v = sum(spend(r) for r in campaign_rows)
    total_conv_v  = sum(conversions(r) for r in campaign_rows)
    best_roas_v   = max((roas(r) for r in campaign_rows if roas(r)), default=None)
    best_cpl_v    = min((cpl(r)  for r in campaign_rows if cpl(r)),  default=None)
    n_green  = sum(1 for _, t, _ in classified if t == "green")
    n_amber  = sum(1 for _, t, _ in classified if t == "yellow")
    n_red_ct = sum(1 for _, t, _ in classified if t == "red")
    n_grey   = sum(1 for _, t, _ in classified if t == "grey")

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.create_sheet("📊 Summary")
    ws.merge_cells("A1:D1")
    c = ws["A1"]; c.value = f"{acct_name} — Google Ads Performance Report"
    c.font = _font(bold=True, size=14); c.alignment = _align("left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Last {days} days  ·  Generated: {generated_at.strftime('%d %b %Y %H:%M')}  ·  {CONFIG['brand']} Performance Marketing"
    c.font = _font(size=10, color="6B7280", italic=True)

    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = f"Account: {acct_name} ({acct_id})  |  Currency: {currency}  |  API: Google Ads {CONFIG['api_version']}"
    c.font = _font(size=10, color="6B7280")

    ws.merge_cells("A5:D5")
    c = ws["A5"]; c.value = "ACCOUNT KPIs"
    c.fill = _fill(NAVY); c.font = _font(bold=True, size=10, color=WHITE); c.alignment = _align("left")

    kpi_rows = [
        ("Total Ad Spend",        f"₹{int(total_spend_v):,}",              None),
        ("Total Conversions",     f"{int(total_conv_v):,}",                 None),
        ("Best CPL",              f"₹{int(best_cpl_v):,}" if best_cpl_v else "N/A",
         _fill(C_GREEN) if best_cpl_v and best_cpl_v <= T["cpl_green"] else None),
        ("Best ROAS",             f"{best_roas_v:.2f}×" if best_roas_v else "N/A",
         _fill(C_GREEN) if best_roas_v and best_roas_v >= T["roas_green"] else None),
        ("Active Campaigns",      len(campaign_rows),                       None),
        ("Active Ad Groups",      len(adgroup_rows),                        None),
    ]
    for i, (label, value, cell_fill) in enumerate(kpi_rows, 6):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, size=10)
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = _font(size=10)
        if cell_fill: vc.fill = cell_fill
        ws.row_dimensions[i].height = 17

    ws.merge_cells("A14:D14")
    c = ws["A14"]; c.value = "CAMPAIGN HEALTH BREAKDOWN"
    c.fill = _fill(NAVY); c.font = _font(bold=True, size=10, color=WHITE); c.alignment = _align("left")

    for i, (label, count, hfill) in enumerate([
        ("🟢 Performing",      n_green,  _fill(C_GREEN)),
        ("🟡 Watch",           n_amber,  _fill(C_AMBER)),
        ("🔴 Underperforming", n_red_ct, _fill(C_RED)),
        ("⚪ Too Early",        n_grey,   _fill(C_GREY)),
    ], 15):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _font(bold=True, size=10)
        if hfill: lc.fill = hfill
        ws.cell(row=i, column=2, value=count).font = _font(size=10)
        ws.row_dimensions[i].height = 17

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2: Campaigns ───────────────────────────────────
    ws2 = wb.create_sheet("📋 Campaigns")
    cols = [
        ("Campaign", 40), ("Spend (₹)", 14), ("Impressions", 14), ("Clicks", 10),
        ("CTR %", 10), ("Avg CPC (₹)", 12), ("CPM (₹)", 10),
        ("Conversions", 14), ("Conv. Value (₹)", 16), ("CPL (₹)", 12),
        ("ROAS", 10), ("Imp. Share %", 14), ("Status", 18),
    ]
    _write_header(ws2, cols)

    for ri, (r, tier, label) in enumerate(classified, 2):
        cp = cpl(r); ro = roas(r); ish = imp_share(r)
        row_data = [
            r.get("campaign_name", "—"),
            _n(spend(r)), _n(impressions(r)), _n(clicks(r)),
            _n(ctr(r), 2), _n(avg_cpc(r), 2), _n(cpm(r), 2),
            _n(conversions(r)), _n(conv_value(r)),
            _n(cp, 0), _n(ro, 2), _n(ish, 1), label,
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10)
        tf = _tier_fill(tier)
        if tf:
            for ci in range(1, len(cols) + 1):
                ws2.cell(row=ri, column=ci).fill = tf
        if cp:
            ws2.cell(row=ri, column=10).fill = _cpl_fill(cp) or PatternFill()
        ws2.row_dimensions[ri].height = 17

    # ── Sheet 3: Ad Groups ───────────────────────────────────
    ws3 = wb.create_sheet("🗺 Ad Groups")
    cols3 = [
        ("Campaign", 36), ("Ad Group", 36), ("Spend (₹)", 14),
        ("Impressions", 14), ("Clicks", 10), ("CTR %", 10),
        ("Conversions", 14), ("CPL (₹)", 12), ("ROAS", 10),
        ("Imp. Share %", 14), ("Status", 18),
    ]
    _write_header(ws3, cols3)

    for ri, a in enumerate(adgroup_rows, 2):
        cp = cpl(a); ro = roas(a)
        if cp and cp > T["cpl_red"]:             status = "Cap Spend"
        elif cp and cp <= T["cpl_green"] and ctr(a) >= T["ctr_green"]: status = "Scale"
        elif imp_share(a) > 0 and imp_share(a) < T["imp_share_red"]:  status = "Watch"
        else:                                                           status = "Hold"
        row_data = [
            a.get("campaign_name", "—"), a.get("ad_group_name", "—"),
            _n(spend(a)), _n(impressions(a)), _n(clicks(a)), _n(ctr(a), 2),
            _n(conversions(a)), _n(cp, 0), _n(ro, 2), _n(imp_share(a), 1), status,
        ]
        for ci, val in enumerate(row_data, 1):
            ws3.cell(row=ri, column=ci, value=val).font = _font(size=10)
        ws3.row_dimensions[ri].height = 17

    # ── Sheet 4: Ads ─────────────────────────────────────────
    ws4 = wb.create_sheet("🎯 Ads")
    cols4 = [
        ("Campaign", 36), ("Ad Group", 28), ("Ad Name", 40), ("Ad Type", 20),
        ("Spend (₹)", 14), ("Impressions", 14), ("CTR %", 10),
        ("Conversions", 14), ("CPL (₹)", 12), ("Verdict", 18),
    ]
    _write_header(ws4, cols4)

    ri = 2
    for cid, ad_rows in ad_rows_map.items():
        for a in ad_rows:
            cp = cpl(a)
            ad_cpls = {x.get("ad_name"): cpl(x) for x in ad_rows if cpl(x)}
            best_ad = min(ad_cpls, key=ad_cpls.get) if ad_cpls else None
            if a.get("ad_name") == best_ad:          verdict = "Keep + Scale ⭐"
            elif cp and cp > T["cpl_amber"]:         verdict = "Pause ❌"
            else:                                    verdict = "Test further 🔍"
            row_data = [
                a.get("campaign_name", "—"), a.get("ad_group_name", "—"),
                a.get("ad_name", "—"), a.get("ad_type", "—"),
                _n(spend(a)), _n(impressions(a)), _n(ctr(a), 2),
                _n(conversions(a)), _n(cp, 0), verdict,
            ]
            for ci, val in enumerate(row_data, 1):
                ws4.cell(row=ri, column=ci, value=val).font = _font(size=10)
            ws4.row_dimensions[ri].height = 17
            ri += 1

    # ── Sheet 5: Quick Wins ──────────────────────────────────
    ws5 = wb.create_sheet("⚡ Quick Wins")
    ws5.merge_cells("A1:D1")
    c = ws5["A1"]; c.value = "Quick Wins — Automated Recommendations"
    c.fill = _fill(NAVY); c.font = _font(bold=True, size=12, color=WHITE)
    c.alignment = _align("left")
    ws5.row_dimensions[1].height = 26
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 12
    ws5.column_dimensions["C"].width = 44
    ws5.column_dimensions["D"].width = 80

    wins = auto_quick_wins(campaign_rows, adgroup_rows, ad_rows_map, red_rows)
    for ri, w in enumerate(wins, 2):
        ws5.cell(row=ri, column=1, value=w["rank"]).font = _font(bold=True, size=11)
        ws5.cell(row=ri, column=2, value=w.get("label", "")).font = _font(size=10, color="6B7280")
        ws5.cell(row=ri, column=3, value=w["title"]).font = _font(bold=True, size=10)
        bd = ws5.cell(row=ri, column=4, value=w["body"].replace("<strong>", "").replace("</strong>", ""))
        bd.font = _font(size=10); bd.alignment = _align("left", wrap=True)
        ws5.row_dimensions[ri].height = 40

    if hasattr(out_path, "write"):
        wb.save(out_path)
    else:
        wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════
def _parse_args():
    p = argparse.ArgumentParser(description="myHQ Google Ads Analysis Agent")
    p.add_argument("--developer-token",  default=os.getenv("GOOGLE_ADS_DEVELOPER_TOKEN"))
    p.add_argument("--client-id",        default=os.getenv("GOOGLE_ADS_CLIENT_ID"))
    p.add_argument("--client-secret",    default=os.getenv("GOOGLE_ADS_CLIENT_SECRET"))
    p.add_argument("--refresh-token",    default=os.getenv("GOOGLE_ADS_REFRESH_TOKEN"))
    p.add_argument("--customer-id",      default=os.getenv("GOOGLE_ADS_CUSTOMER_ID"))
    p.add_argument("--days",             type=int, default=CONFIG["default_days"])
    p.add_argument("--out",              default=None)
    return p.parse_args()


def main():
    args = _parse_args()

    missing = [k for k, v in {
        "developer-token": args.developer_token,
        "client-id":       args.client_id,
        "client-secret":   args.client_secret,
        "refresh-token":   args.refresh_token,
        "customer-id":     args.customer_id,
    }.items() if not v]
    if missing:
        print(f"\n  ERROR: Missing required arguments: {', '.join(missing)}")
        print("  Provide via flags or env vars (GOOGLE_ADS_DEVELOPER_TOKEN etc.)\n")
        sys.exit(1)

    _log(f"\n  myHQ Google Ads Analysis Agent · {CONFIG['api_version']}")
    _log(f"  Customer ID: {args.customer_id} · Last {args.days} days\n")

    client = GoogleAdsClient(
        args.developer_token, args.client_id,
        args.client_secret,   args.refresh_token,
        args.customer_id,
    )

    _log("  Fetching account info …")
    account = client.account_info()
    _ok(f"{account['name']} ({account['id']})")

    _log("  Fetching campaign insights …")
    campaign_rows = client.campaign_insights(args.days)
    _ok(f"{len(campaign_rows)} campaigns")

    _log("  Fetching ad group insights …")
    adgroup_rows = client.adgroup_insights(args.days)
    _ok(f"{len(adgroup_rows)} ad groups")

    classified = [(r, *classify(r)) for r in campaign_rows]
    red_cids   = [r.get("campaign_id") for r, t, _ in classified if t == "red"]
    ad_rows_map = {}

    if red_cids:
        _log(f"  Fetching ad-level data for {len(red_cids)} underperforming campaign(s) …")
        ad_rows_map = {cid: [] for cid in red_cids}
        all_ads = client.ad_insights(red_cids, args.days)
        for a in all_ads:
            cid = a.get("campaign_id")
            if cid in ad_rows_map:
                ad_rows_map[cid].append(a)
        _ok(f"{sum(len(v) for v in ad_rows_map.values())} ad rows")

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    html_path = args.out or f"Google_Ads_Report_{date_str}.html"
    xlsx_path = html_path.replace(".html", ".xlsx")

    _log("  Generating HTML report …")
    html = generate_html(account, campaign_rows, adgroup_rows, ad_rows_map, args.days, now)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    _ok(html_path)

    _log("  Generating Excel workbook …")
    generate_excel(account, campaign_rows, adgroup_rows, ad_rows_map, args.days, now, xlsx_path)
    _ok(xlsx_path)

    _log(f"\n  Done. Opening {html_path} …\n")
    webbrowser.open(html_path)


if __name__ == "__main__":
    main()
